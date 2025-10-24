"""Spreadsheet parsing utilities."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Optional


@dataclass
class SpreadsheetRow:
    sku: str
    product_name: str
    brand: str
    website: str
    product_url: str
    status: str
    extra: Dict[str, str]

    @property
    def normalized_sku(self) -> str:
        """Return a SKU formatted for image naming."""

        if self.sku.endswith("-p"):
            return self.sku[:-2]
        return self.sku


class SpreadsheetLoader:
    """Load spreadsheet data from CSV or XLSX files."""

    def __init__(self, expected_columns: Optional[Iterable[str]] = None) -> None:
        self.expected_columns = list(expected_columns or [])

    def load(self, path: Path) -> List[SpreadsheetRow]:
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(path)

        if path.suffix.lower() == ".csv":
            rows = list(self._load_csv(path))
        elif path.suffix.lower() in {".xlsx", ".xlsm"}:
            rows = list(self._load_xlsx(path))
        else:
            raise ValueError(f"Unsupported spreadsheet format: {path.suffix}")

        if self.expected_columns and rows:
            missing = set(self.expected_columns) - set(rows[0].extra.keys()) - {
                "sku",
                "product_name",
                "brand",
                "website",
                "product_url",
                "status",
            }
            if missing:
                raise ValueError(f"Spreadsheet missing expected columns: {sorted(missing)}")
        return rows

    def _load_csv(self, path: Path) -> Iterator[SpreadsheetRow]:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            yield from self._rows_from_dicts(reader)

    def _load_xlsx(self, path: Path) -> Iterator[SpreadsheetRow]:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:  # pragma: no cover - optional dependency
            raise ImportError("openpyxl is required to read Excel files") from exc

        workbook = load_workbook(path, read_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {header: (value if value is not None else "") for header, value in zip(headers, row)}
            yield from self._rows_from_dicts([row_dict])

    def _rows_from_dicts(self, rows: Iterable[Dict[str, str]]) -> Iterator[SpreadsheetRow]:
        for row in rows:
            normalized = {key.strip().lower(): (value.strip() if isinstance(value, str) else value) for key, value in row.items() if key}
            base_kwargs = {
                "sku": str(normalized.get("sku", "")).strip(),
                "product_name": str(normalized.get("product_name", "")).strip(),
                "brand": str(normalized.get("brand", "")).strip(),
                "website": str(normalized.get("website", "")).strip(),
                "product_url": str(normalized.get("product_url", "")).strip(),
                "status": str(normalized.get("status", "")).strip().lower(),
            }
            extra = {
                key: value
                for key, value in normalized.items()
                if key not in base_kwargs or key in {"price", "availability"}
            }
            yield SpreadsheetRow(extra=extra, **base_kwargs)
